#!/usr/bin/env python3
#------------------------------------------------------------------------------
# Author: 	Joel Gurnett
# Desc: 	find standings for keywords
# Date:		October 18, 2020
#------------------------------------------------------------------------------
import getopt, sys
from googlesearch import search
from openpyxl import load_workbook
from datetime import date

def main(argv):
	inputfile = ""
	site = ""
	standing = {}

	# gets input from user
	try: 
		opts, args = getopt.getopt(argv, "hi:", ["ifile="])
	except getopt.GetoptError:
		sys.exit(2)
	for opt, arg in opts:
		# display how to use the program
		if opt == '-h':
			print('you have reached the help menu:')
			print('\t-h\tDisplay This menu')
			print('\t-i\tSpecify file for input\n')
			print('\tUsage: ./keywordSearch.py -i excel.xlsx')
			sys.exit()

		# display inputfile and save for later
		elif opt in ('-i', '--ifile'):
			inputfile = arg
			print('input file is: ' + inputfile)
	
	# find positions and write them to an excel file
	try:
		# load up workbook
		workbook = load_workbook(filename=inputfile)
		sheet = workbook.active
		site = sheet.title
		countryCode = sheet.cell(row=1, column=1).value
		
		query = ""
		file = open(inputfile, 'r')
		row = 2
		col = 2

		# find next empty column
		while sheet.cell(row=1, column=col).value != None:		
			col = col + 1
		# add current date to first row
		sheet.cell(row=row-1, column=col).value = date.today().strftime("%Y-%m-%d") 
		
		query = sheet.cell(row=row, column=1).value

		# iterate through all keywords
		while query != None:
			if query != '-':
				numPage = 10
				print("searching for: " + query + "...")
				try:
					# stop: search 10 pages 
					# countryCode: which contry to search
					# num: 10 results per page 
					# pause: wait 2 seconds every 10 results to prevent getting banned
					searchResults = search(query, stop=100, country=countryCode, num=10, pause=2.0)

					# find position of our keyword
					for i, result in enumerate(searchResults):
						count = i + 1
						try:
							if site in result:
								page = 1
								position = count
								if count > 9:
									page = int(count / 10) + 1 
									position = count % 10

								# add postition to excel sheet
								sheet.cell(row=row, column=col).value = count
								standing[query] = "Page: " + str(page) + " position: " + str(position)

								# display standing in terminal window
								print(standing[query] + "\n")
								break
						except:
							print("error!")

					# display not found if keyword isn't in top 10 pages
					if query not in standing:
						sheet.cell(row=row, column=col).value = 'Not Found'
						print('Query: ' + query + " not found!\n")
				except:
					print("\nYou have made too many requests!!!")
					print("Saving current requests...")
					workbook.save(filename=inputfile)
					time.sleep(3)
					print("\nResults saved! please try again later")
					quit()
			row = row + 1

			# get next keyword
			query = sheet.cell(row=row, column=1).value

		# save excel workbook
		workbook.save(filename=inputfile)

	# display if the file doesn't exist
	except FileNotFoundError:
		print("The file you entered doesn't exist")

if __name__ == "__main__":
	main(sys.argv[1:])
